import io

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

SUMMARY_LABELS = [
    ("total", "Jami buyurtmalar"),
    ("completed", "Tugallangan"),
    ("in_progress", "Jarayonda"),
    ("new", "Yangi"),
    ("cancelled", "Bekor qilingan"),
    ("overdue", "Kechikkan"),
]

HEADER_FILL = "#6B4423"
DANGER_FILL = "#C0392B"
GRID_COLOR = "#D8C6AC"
ROW_ALT = "#F7F1E8"


def _autosize(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 40)


def render_orders_report_excel(data):
    wb = Workbook()

    ws = wb.active
    ws.title = "Umumiy"
    ws.append(["Ko'rsatkich", "Qiymat"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for key, label in SUMMARY_LABELS:
        ws.append([label, data["summary"][key]])

    ws2 = wb.create_sheet("Ishchilar samaradorligi")
    ws2.append(["F.I.Sh.", "Roli", "Tugatgan", "Jarayonda", "Oxirgi tugatgan sana"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for worker in data["worker_performance"]:
        ws2.append([
            worker["name"], worker["role"], worker["completed_count"], worker["in_progress_count"],
            worker["last_completed_at"] or "-",
        ])

    ws3 = wb.create_sheet("Kechikkan buyurtmalar")
    ws3.append(["Buyurtma", "Mijoz", "Muddat", "Necha kun kechikkan", "Status"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for order in data["overdue_orders"]:
        ws3.append([order["order_no"], order["customer_name"], order["deadline"], order["days_overdue"], order["status_label"]])

    for sheet in wb.worksheets:
        _autosize(sheet)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _styled_table(rows, header_fill, repeat_header=True):
    table = Table(rows, repeatRows=1 if repeat_header else 0)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_fill)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(GRID_COLOR)),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(ROW_ALT)]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def render_orders_report_pdf(data):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Hisobot")
    styles = getSampleStyleSheet()
    summary = data["summary"]

    summary_rows = [[label for _, label in SUMMARY_LABELS], [summary[key] for key, _ in SUMMARY_LABELS]]
    summary_table = _styled_table(summary_rows, HEADER_FILL, repeat_header=False)

    worker_rows = [["F.I.Sh.", "Roli", "Tugatgan", "Jarayonda", "Oxirgi tugatgan"]]
    for worker in data["worker_performance"]:
        worker_rows.append([
            worker["name"], worker["role"], worker["completed_count"], worker["in_progress_count"],
            (worker["last_completed_at"] or "-")[:10],
        ])
    worker_table = _styled_table(worker_rows, HEADER_FILL)

    story = [
        Paragraph("Buyurtmalar va ishchilar hisoboti", styles["Title"]),
        Spacer(1, 10),
        summary_table,
        Spacer(1, 18),
        Paragraph("Ishchilar samaradorligi", styles["Heading2"]),
        Spacer(1, 6),
        worker_table,
        Spacer(1, 18),
        Paragraph("Kechikkan buyurtmalar", styles["Heading2"]),
        Spacer(1, 6),
    ]

    if data["overdue_orders"]:
        overdue_rows = [["Buyurtma", "Mijoz", "Muddat", "Kechikish (kun)", "Status"]]
        for order in data["overdue_orders"]:
            overdue_rows.append([order["order_no"], order["customer_name"], order["deadline"], order["days_overdue"], order["status_label"]])
        story.append(_styled_table(overdue_rows, DANGER_FILL))
    else:
        story.append(Paragraph("Kechikkan buyurtmalar yo'q", styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    return buf
